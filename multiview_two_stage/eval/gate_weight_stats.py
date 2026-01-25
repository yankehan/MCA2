
import argparse
import os
import random
import sys
import time
from datetime import datetime
import numpy as np
import torch
import torch.backends.cudnn as cudnn
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    _HAS_OPENPYXL = True
except ImportError:
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    _HAS_OPENPYXL = False
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
sys.path.insert(0, project_root)
os.chdir(project_root)
print(f"✓ 工作目录已切换到: {os.getcwd()}")
from multiview_two_stage.dataset import load_all_views_full
from multiview_two_stage.model import MultiViewContrastiveModel
from multiview_two_stage.trainer import SimpleMultiViewTrainer
try:
    from multiview_two_stage.eval.dataset_configs import DATASET_CONFIGS, print_dataset_config
    USE_DATASET_CONFIGS = True
except ImportError:
    USE_DATASET_CONFIGS = False
    print("⚠️  未找到dataset_configs.py，将使用默认配置")
DEFAULT_DATASETS = ','.join(list(DATASET_CONFIGS.keys())) if USE_DATASET_CONFIGS else 'olid'
VIEW_COMBINATIONS = {
    "O-ada+small+large": ['openai_ada', 'openai_small', 'openai_large'],
    "multi-view": ['openai_large', 'bert', 'qwen', 'llama'],
}
MODEL_CONFIG = {
    'latent_dim': 128,
    'hidden_dims': [512, 256],
    'activation': 'relu',
    'batchnorm': True,
}
DATA_NAME = 'olid'
TARGET_COMBO = 'all'
STAT_SPLIT = 'test'
EVAL_BATCH_SIZE = None
NUM_PRINT = 20
PRINT_SEED = 123
VIEW_GATE_HIDDEN_DIMS = None
VIEW_GATE_TEMPERATURE = 2.0
EXCEL_PATH = None
EXCEL_SHEET = 'gate_weight_stats'
def _default_excel_path_for_dataset(dataset_name: str):
    name = str(dataset_name).strip() if dataset_name is not None else 'dataset'
    if not name:
        name = 'dataset'
    eval_dir = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(eval_dir, f"gate_weight_stats_{name}.xlsx")
TRAIN_CONFIG = {
    'lambda_recon': 1.0,
    'lambda_contrastive': 1.0,
    'temperature': 0.5,
    'batch_size': None,
    'score_weight_recon': 0.3,
    'score_weight_consistency': 0.4,
}
BASE_TRAIN_CONFIG = TRAIN_CONFIG.copy()
NUM_EPOCHS = 200
LEARNING_RATE = 0.002
GATE_LEARNING_RATE = 0.001
PRINT_EVERY = 1
BASE_NUM_EPOCHS = NUM_EPOCHS
BASE_LEARNING_RATE = LEARNING_RATE
FIRST_STAGE_EPOCHS = 50
SECOND_STAGE_EPOCHS = 1
BASE_FIRST_STAGE_EPOCHS = FIRST_STAGE_EPOCHS
BASE_SECOND_STAGE_EPOCHS = SECOND_STAGE_EPOCHS
BASE_GATE_LEARNING_RATE = GATE_LEARNING_RATE
def _print_kv_block(title: str, items):
    print("\n--- " + str(title) + " ---")
    for k, v in items:
        print(f"{k}: {v}")
def _format_ratio(count: int, total: int):
    total = max(1, int(total))
    return f"{count} / {total} ({count / total * 100:.2f}%)"
def _print_view_table(view_names, mean_per_view, var_per_view, argmax_counts, n_samples):
    rows = []
    for i, v in enumerate(view_names):
        c = int(argmax_counts.get(v, 0))
        rows.append((
            str(v),
            f"{float(mean_per_view[i]):.6f}",
            f"{float(var_per_view[i]):.6f}",
            _format_ratio(c, n_samples),
        ))
    col_names = ("View", "Mean", "Var", "Argmax")
    widths = [len(c) for c in col_names]
    for r in rows:
        for j in range(len(widths)):
            widths[j] = max(widths[j], len(str(r[j])))
    header = " | ".join(col_names[j].ljust(widths[j]) for j in range(len(widths)))
    sep = "-+-".join("-" * widths[j] for j in range(len(widths)))
    print(header)
    print(sep)
    for r in rows:
        print(" | ".join(str(r[j]).ljust(widths[j]) for j in range(len(widths))))
def _build_gate_stats_text(entropy_mean, view_names, mean_per_view, var_per_view, argmax_counts, n_samples):
    rows = []
    for i, v in enumerate(view_names):
        c = int(argmax_counts.get(v, 0))
        rows.append((
            str(v),
            f"{float(mean_per_view[i]):.6f}",
            f"{float(var_per_view[i]):.6f}",
            _format_ratio(c, n_samples),
        ))
    col_names = ("View", "Mean", "Var", "Argmax")
    widths = [len(c) for c in col_names]
    for r in rows:
        for j in range(len(widths)):
            widths[j] = max(widths[j], len(str(r[j])))
    header = " | ".join(col_names[j].ljust(widths[j]) for j in range(len(widths)))
    sep = "-+-".join("-" * widths[j] for j in range(len(widths)))
    lines = [
        "Gate 权重统计".center(80, "-"),
        f"Gate entropy mean: {float(entropy_mean):.6f}",
        header,
        sep,
    ]
    for r in rows:
        lines.append(" | ".join(str(r[j]).ljust(widths[j]) for j in range(len(widths))))
    return "\n".join(lines)
def _append_rows_to_excel(excel_path, sheet_name, title, header, rows):
    if not _HAS_OPENPYXL:
        raise RuntimeError("未安装openpyxl，无法保存Excel。请先安装: pip install openpyxl")
    excel_path_abs = os.path.abspath(excel_path)
    os.makedirs(os.path.dirname(excel_path_abs) or '.', exist_ok=True)
    if os.path.exists(excel_path_abs):
        wb = load_workbook(excel_path_abs)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    n_cols = len(header)
    first_row = [c.value for c in ws[1][:n_cols]]
    second_row = [c.value for c in ws[2][:n_cols]] if ws.max_row >= 2 else [None] * n_cols
    is_empty = ws.max_row == 1 and all(v is None for v in ws[1])
    has_title_and_header = (ws.max_row >= 2 and ws.cell(row=1, column=1).value is not None and second_row == list(header))
    is_old_header_only = (first_row == list(header))
    if is_empty:
        is_old_header_only = False
        has_title_and_header = False
    if is_old_header_only:
        ws.insert_rows(1)
        first_row = [c.value for c in ws[1][:n_cols]]
        second_row = [c.value for c in ws[2][:n_cols]] if ws.max_row >= 2 else [None] * n_cols
        has_title_and_header = (ws.max_row >= 2 and ws.cell(row=1, column=1).value is not None and second_row == list(header))
    if not has_title_and_header:
        ws.cell(row=1, column=1).value = str(title)
        if n_cols >= 2:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        if Font is not None and Alignment is not None:
            ws.cell(row=1, column=1).font = Font(bold=True)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        for j, col_name in enumerate(header, start=1):
            ws.cell(row=2, column=j).value = col_name
            if Font is not None:
                ws.cell(row=2, column=j).font = Font(bold=True)
    for r in rows:
        ws.append(list(r))
    wb.save(excel_path_abs)
    return excel_path_abs
def _save_gate_stats_to_excel(
    excel_path,
    sheet_name,
    dataset_name,
    combo_name,
    split_name,
    seed,
    entropy_mean,
    view_names,
    mean_per_view,
    var_per_view,
    argmax_counts,
    n_samples,
):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    header = (
        'time', 'dataset', 'combo', 'split', 'seed', 'entropy_mean',
        'view', 'mean', 'var', 'argmax_count', 'argmax_total', 'argmax_pct'
    )
    rows = []
    for i, v in enumerate(view_names):
        c = int(argmax_counts.get(v, 0))
        pct = float(c) / float(max(1, int(n_samples)))
        rows.append((
            ts,
            str(dataset_name),
            str(combo_name),
            str(split_name),
            int(seed),
            float(entropy_mean),
            str(v),
            float(mean_per_view[i]),
            float(var_per_view[i]),
            int(c),
            int(n_samples),
            float(pct),
        ))
    try:
        saved_path = _append_rows_to_excel(excel_path, sheet_name, 'Gate 权重统计结果', header, rows)
        raw_sheet = f"{sheet_name}_raw"
        raw_header = ('time', 'dataset', 'combo', 'split', 'seed', 'text')
        raw_text = _build_gate_stats_text(entropy_mean, view_names, mean_per_view, var_per_view, argmax_counts, n_samples)
        _append_rows_to_excel(excel_path, raw_sheet, 'Gate 权重统计文本记录', raw_header, [(ts, str(dataset_name), str(combo_name), str(split_name), int(seed), raw_text)])
        print(
            f"✓ Excel已追加保存: {saved_path} | sheet={sheet_name} | dataset={dataset_name} | combo={combo_name} | split={split_name}"
        )
    except Exception as e:
        print(
            f"✗ Excel保存失败: path={os.path.abspath(excel_path)} | dataset={dataset_name} | combo={combo_name} | err={type(e).__name__}: {e}"
        )
def _set_requires_grad(module, requires_grad: bool):
    for p in module.parameters():
        p.requires_grad = bool(requires_grad)
def set_random_seed(seed: int):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    cudnn.deterministic = True
    cudnn.benchmark = False
    print(f"随机数种子已设置为: {seed}")
def _parse_int_list(s):
    if s is None:
        return None
    s = s.strip()
    if not s:
        return None
    return [int(x) for x in s.split(',') if x.strip()]
def _compute_weights_in_batches(model, views_dict, batch_size=None):
    model.eval()
    n = list(views_dict.values())[0].shape[0]
    if batch_size is None or batch_size >= n:
        out = model.compute_view_weights(views_dict)
        if out is None:
            raise RuntimeError("当前模型没有 view_gate，无法计算 gate 权重")
        weights, _ = out
        return weights
    all_weights = []
    with torch.no_grad():
        for i in range(0, n, batch_size):
            end = min(i + batch_size, n)
            batch_views = {k: v[i:end] for k, v in views_dict.items()}
            out = model.compute_view_weights(batch_views)
            if out is None:
                raise RuntimeError("当前模型没有 view_gate，无法计算 gate 权重")
            batch_weights, _ = out
            all_weights.append(batch_weights)
    return torch.cat(all_weights, dim=0)
def _train_two_stage(model, train_views_dict, eval_views_dict, eval_labels, device):
    if hasattr(model, 'set_view_gate_mode'):
        model.set_view_gate_mode('uniform')
    if hasattr(model, 'view_gate') and model.view_gate is not None and hasattr(model.view_gate, 'reset_to_uniform'):
        model.view_gate.reset_to_uniform()
    if hasattr(model, 'view_gate') and model.view_gate is not None:
        _set_requires_grad(model.view_gate, False)
    stage1_params = [p for n, p in model.named_parameters() if not n.startswith('view_gate.')]
    optimizer_stage1 = torch.optim.Adam(stage1_params, lr=LEARNING_RATE)
    trainer_stage1 = SimpleMultiViewTrainer(
        model=model,
        optimizer=optimizer_stage1,
        device=device,
        config=TRAIN_CONFIG,
    )
    print("\n" + "Stage 1: 训练 AE + 对比学习（Gate 固定为均匀，不更新）".center(80, "-"))
    print(f"first_stage_epochs={FIRST_STAGE_EPOCHS}")
    stage1_start = time.time()
    for epoch in range(int(FIRST_STAGE_EPOCHS)):
        loss_dict = trainer_stage1.train_epoch(train_views_dict)
        if (epoch + 1) % PRINT_EVERY == 0:
            elapsed = time.time() - stage1_start
            print(
                f"[Stage1] Epoch {epoch + 1}/{FIRST_STAGE_EPOCHS} | "
                f"Loss: {loss_dict['total']:.4f} | Recon: {loss_dict['recon']:.4f} | Contr: {loss_dict['contrastive']:.4f} | "
                f"Time: {elapsed:.1f}s"
            )
    stage1_time = time.time() - stage1_start
    auc_off, auprc_off, _ = trainer_stage1.evaluate(eval_views_dict, eval_labels)
    print(f"\n[Stage1] gate_off AUC: {float(auc_off):.4f} AUPRC: {float(auprc_off):.4f}")
    if hasattr(model, 'set_view_gate_mode'):
        model.set_view_gate_mode('learned')
    for n, p in model.named_parameters():
        if n.startswith('view_gate.'):
            p.requires_grad = True
        else:
            p.requires_grad = False
    if hasattr(model, 'view_gate') and model.view_gate is not None and hasattr(model.view_gate, 'reset_to_uniform'):
        model.view_gate.reset_to_uniform()
    gate_params = [p for n, p in model.named_parameters() if n.startswith('view_gate.')]
    optimizer_stage2 = torch.optim.Adam(gate_params, lr=float(GATE_LEARNING_RATE))
    stage2_config = dict(TRAIN_CONFIG)
    stage2_config['train_mode'] = 'gate_only'
    trainer_stage2 = SimpleMultiViewTrainer(
        model=model,
        optimizer=optimizer_stage2,
        device=device,
        config=stage2_config,
    )
    print("\n" + "Stage 2: 冻结 AE+CL，仅训练 Gate（每 epoch 评估 AUC）".center(80, "-"))
    print(f"second_stage_epochs={SECOND_STAGE_EPOCHS}, gate_lr={float(GATE_LEARNING_RATE)}")
    best_auc = -1.0
    best_auprc = -1.0
    best_state = None
    stage2_start = time.time()
    for epoch in range(int(SECOND_STAGE_EPOCHS)):
        loss_dict = trainer_stage2.train_epoch(train_views_dict)
        auc, auprc, _ = trainer_stage2.evaluate(eval_views_dict, eval_labels)
        elapsed = time.time() - stage2_start
        print(
            f"[Stage2] Epoch {epoch + 1}/{SECOND_STAGE_EPOCHS} | "
            f"Loss: {loss_dict['total']:.4f} | Recon: {loss_dict['recon']:.4f} | Contr: {loss_dict['contrastive']:.4f} | "
            f"Test AUC: {float(auc):.4f} AUPRC: {float(auprc):.4f} | Time: {elapsed:.1f}s"
        )
        if auc > best_auc:
            best_auc = float(auc)
            best_auprc = float(auprc)
            best_state = {k: v.detach().clone() for k, v in model.state_dict().items()}
    if best_state is not None:
        model.load_state_dict(best_state)
    stage2_time = time.time() - stage2_start
    print(f"\n[Stage2] gate_on best AUC: {float(best_auc):.4f} AUPRC: {float(best_auprc):.4f}")
    del trainer_stage1
    del trainer_stage2
    del optimizer_stage1
    del optimizer_stage2
    return float(auc_off), float(best_auc), float(auprc_off), float(best_auprc), float(stage1_time), float(stage2_time)
def _run_one_combo(combo_name, seed, device):
    view_names = VIEW_COMBINATIONS[combo_name]
    print("\n" + "=" * 80)
    print("Gate 权重统计 (两阶段训练后统计 gate 输出分布)")
    print("=" * 80)
    print(f"\n数据集: {DATA_NAME}")
    print(f"组合: {combo_name}")
    print(f"视图: {view_names}")
    print(f"统计 split: {STAT_SPLIT}")
    _print_kv_block(
        "训练/评估配置",
        [
            ("lr", LEARNING_RATE),
            ("first_stage_epochs", FIRST_STAGE_EPOCHS),
            ("second_stage_epochs", SECOND_STAGE_EPOCHS),
            ("gate_lr", GATE_LEARNING_RATE),
            ("batch_size", TRAIN_CONFIG.get('batch_size', None)),
            ("lambda_recon", TRAIN_CONFIG['lambda_recon']),
            ("lambda_contrastive", TRAIN_CONFIG['lambda_contrastive']),
            ("temperature", TRAIN_CONFIG.get('temperature', 0.5)),
            ("score_weight_recon", TRAIN_CONFIG.get('score_weight_recon', 0.3)),
            ("score_weight_consistency", TRAIN_CONFIG.get('score_weight_consistency', 0.4)),
            ("view_gate_temperature", VIEW_GATE_TEMPERATURE),
        ],
    )
    print("\n加载训练集数据（只包含正常数据）...")
    train_views_dict, _, _ = load_all_views_full(
        data_name=DATA_NAME,
        view_names=view_names,
        split='train',
        embeddings_dir='embeddings',
        normalize=True,
        device=device,
    )
    print("\n加载统计<测试集>数据...")
    stat_views_dict, stat_labels, _ = load_all_views_full(
        data_name=DATA_NAME,
        view_names=view_names,
        split=STAT_SPLIT,
        embeddings_dir='embeddings',
        normalize=True,
        device=device,
    )
    view_dims = {name: emb.shape[1] for name, emb in train_views_dict.items()}
    print(f"\n视图维度: {view_dims}")
    model = MultiViewContrastiveModel(
        view_dims=view_dims,
        latent_dim=MODEL_CONFIG['latent_dim'],
        hidden_dims=MODEL_CONFIG['hidden_dims'],
        activation=MODEL_CONFIG['activation'],
        batchnorm=MODEL_CONFIG['batchnorm'],
        use_view_gate=True,
        view_gate_hidden_dims=VIEW_GATE_HIDDEN_DIMS,
        view_gate_temperature=VIEW_GATE_TEMPERATURE,
    ).to(device)
    auc_off, auc_on, auprc_off, auprc_on, stage1_time, stage2_time = _train_two_stage(
        model=model,
        train_views_dict=train_views_dict,
        eval_views_dict=stat_views_dict,
        eval_labels=stat_labels,
        device=device,
    )
    print("\n结果:")
    print(f"  gate_off AUC (Stage1): {auc_off:.4f} AUPRC: {auprc_off:.4f} | time: {stage1_time:.1f}s")
    print(f"  gate_on  AUC (Stage2): {auc_on:.4f} AUPRC: {auprc_on:.4f} | time: {stage2_time:.1f}s")
    print(f"  ΔAUC: {auc_on - auc_off:+.4f} ΔAUPRC: {auprc_on - auprc_off:+.4f}")
    combo_auc = auc_on
    with torch.no_grad():
        weights = _compute_weights_in_batches(model, stat_views_dict, batch_size=EVAL_BATCH_SIZE)
    view_order = model.view_names
    if view_order != view_names:
        print(f"\n⚠️  view_order 与输入 view_names 不一致: {view_order} vs {view_names}")
    n = weights.shape[0]
    mean_per_view = weights.mean(dim=0)
    var_per_view = weights.var(dim=0, unbiased=False)
    eps = 1e-12
    entropy = -(weights * (weights + eps).log()).sum(dim=1)
    entropy_mean = entropy.mean().item()
    print("\n" + "Gate 权重统计".center(80, "-"))
    print(f"Gate entropy mean: {entropy_mean:.6f}")
    argmax_idx = weights.argmax(dim=1).cpu().numpy()
    counts = {v: 0 for v in view_order}
    for idx in argmax_idx:
        counts[view_order[int(idx)]] += 1
    _print_view_table(view_order, mean_per_view, var_per_view, counts, n)
    _save_gate_stats_to_excel(
        excel_path=EXCEL_PATH,
        sheet_name=EXCEL_SHEET,
        dataset_name=DATA_NAME,
        combo_name=combo_name,
        split_name=STAT_SPLIT,
        seed=seed,
        entropy_mean=entropy_mean,
        view_names=view_order,
        mean_per_view=mean_per_view.detach().cpu().numpy(),
        var_per_view=var_per_view.detach().cpu().numpy(),
        argmax_counts=counts,
        n_samples=n,
    )
    print("\n" + "随机打印样本 gate 权重".center(80, "-"))
    rng = np.random.default_rng(PRINT_SEED)
    k = min(NUM_PRINT, n)
    sample_indices = rng.choice(n, size=k, replace=False)
    weights_cpu = weights.detach().cpu().numpy()
    labels_cpu = stat_labels.detach().cpu().numpy() if stat_labels is not None else None
    for idx in sample_indices:
        w = weights_cpu[idx]
        parts = [f"{v}={w[i]:.6f}" for i, v in enumerate(view_order)]
        label_str = str(int(labels_cpu[idx])) if labels_cpu is not None else 'NA'
        print(f"idx={int(idx)} label={label_str} | " + ", ".join(parts) + f" | argmax={view_order[int(np.argmax(w))]}")
    return combo_auc
def main(seed=42):
    set_random_seed(seed)
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"设备: {device}")
    print(f"种子: {seed}")
    if TARGET_COMBO == 'all':
        combos_to_run = list(VIEW_COMBINATIONS.keys())
    else:
        combos_to_run = [TARGET_COMBO]
    combo_auc_map = {}
    for combo_name in combos_to_run:
        combo_auc = _run_one_combo(combo_name, seed=seed, device=device)
        combo_auc_map[combo_name] = combo_auc
    print("\n" + "=" * 80)
    print("汇总结果")
    print("=" * 80)
    for combo_name in combos_to_run:
        auc = combo_auc_map.get(combo_name, None)
        auc_str = 'NA' if auc is None else f"{float(auc):.4f}"
        print(f"{combo_name.ljust(20)} | best_auc: {auc_str}")
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Gate 权重统计脚本 (尽量对齐 ourmethod_eval.py)')
    parser.add_argument('--dataset', type=str, default="olid",
                        help='数据集名称，支持逗号分隔多个 (例如: bbc,olid)，可以选DEFAULT_DATASETS表示执行所有')
    parser.add_argument('--seed', type=int, default=42,
                        help='随机数种子 (默认: 42)')
    combo_choices = ['all'] + list(VIEW_COMBINATIONS.keys())
    parser.add_argument('--combo', type=str, default='all', choices=combo_choices,
                        help='视图组合名称；默认 all 表示依次统计所有组合')
    parser.add_argument('--split', type=str, default='test', choices=['train', 'test'],
                        help='统计 gate 权重使用的数据 split (train/test)')
    parser.add_argument('--lambda_recon', type=float, default=1.0,
                        help='重构损失权重 (默认: 1.0)')
    parser.add_argument('--lambda_contrastive', type=float, default=1.0,
                        help='对比学习损失权重 (默认: 1.0)')
    parser.add_argument('--temperature', type=float, default=0.5,
                        help='对比学习温度参数 (默认: 0.5)')
    parser.add_argument('--epochs', type=int, default=None,
                        help='训练轮数 (默认: 自动根据数据集选择)')
    parser.add_argument('--first_stage_epochs', type=int, default=None,
                        help='第一阶段训练轮数 (默认: 自动根据数据集选择；优先级高于 --epochs)')
    parser.add_argument('--second_stage_epochs', type=int, default=BASE_SECOND_STAGE_EPOCHS,
                        help='第二阶段训练轮数 (默认: 30)')
    parser.add_argument('--lr', type=float, default=None,
                        help='学习率 (默认: 自动根据数据集选择)')
    parser.add_argument('--gate_lr', type=float, default=BASE_GATE_LEARNING_RATE,
                        help='第二阶段 gate 学习率 (默认: 0.001)')
    parser.add_argument('--score_weight_recon', type=float, default=None,
                        help='异常分数中重构误差权重 (默认: 自动根据数据集选择)')
    parser.add_argument('--score_weight_consistency', type=float, default=None,
                        help='异常分数中一致性分数权重 (默认: 自动根据数据集选择)')
    parser.add_argument('--batch_size', type=int, default=None,
                        help='Mini-batch大小 (默认: None表示使用全部数据，推荐256/512/1024以避免OOM)')
    parser.add_argument('--view_gate_hidden_dims', type=str, default="256,128",
                        help='gate 的隐藏层维度，逗号分隔，例如 256 或 256,128；为空表示线性 gate')
    parser.add_argument('--view_gate_temperature', type=float, default=1,
                        help='gate softmax 温度参数 (默认: 2.0；设为1.0表示不使用温度缩放)')
    parser.add_argument('--eval_batch_size', type=int, default=None,
                        help='统计 gate 权重时的 batch size（避免一次性前向 OOM）')
    parser.add_argument('--num_print', type=int, default=20,
                        help='随机打印多少个样本的 gate 权重 (默认: 20)')
    parser.add_argument('--print_seed', type=int, default=123,
                        help='随机打印样本的种子 (默认: 123)')
    parser.set_defaults(use_auto_config=True)
    parser.add_argument('--use_auto_config', dest='use_auto_config', action='store_true',
                        help='使用数据集自适应配置（默认: 开启）')
    parser.add_argument('--no_auto_config', dest='use_auto_config', action='store_false',
                        help='禁用数据集自适应配置')
    args = parser.parse_args()
    dataset_list = [d.strip() for d in args.dataset.split(',') if d.strip()]
    for dataset_name in dataset_list:
        DATA_NAME = dataset_name
        TARGET_COMBO = args.combo
        STAT_SPLIT = args.split
        EVAL_BATCH_SIZE = args.eval_batch_size
        NUM_PRINT = args.num_print
        PRINT_SEED = args.print_seed
        VIEW_GATE_HIDDEN_DIMS = _parse_int_list(args.view_gate_hidden_dims)
        VIEW_GATE_TEMPERATURE = float(args.view_gate_temperature)
        EXCEL_PATH = _default_excel_path_for_dataset(dataset_name)
        TRAIN_CONFIG.clear()
        TRAIN_CONFIG.update(BASE_TRAIN_CONFIG)
        NUM_EPOCHS = BASE_NUM_EPOCHS
        LEARNING_RATE = BASE_LEARNING_RATE
        FIRST_STAGE_EPOCHS = BASE_FIRST_STAGE_EPOCHS
        SECOND_STAGE_EPOCHS = BASE_SECOND_STAGE_EPOCHS
        GATE_LEARNING_RATE = BASE_GATE_LEARNING_RATE
        if args.use_auto_config and USE_DATASET_CONFIGS:
            print("\n" + "🔧 使用数据集自适应配置".center(80, "="))
            dataset_config = print_dataset_config(dataset_name)
            first_stage_epochs_arg = args.first_stage_epochs if args.first_stage_epochs is not None else args.epochs
            FIRST_STAGE_EPOCHS = first_stage_epochs_arg if first_stage_epochs_arg is not None else dataset_config['num_epochs']
            LEARNING_RATE = args.lr if args.lr is not None else dataset_config['learning_rate']
            SECOND_STAGE_EPOCHS = int(args.second_stage_epochs)
            GATE_LEARNING_RATE = float(args.gate_lr)
            TRAIN_CONFIG['lambda_recon'] = args.lambda_recon if args.lambda_recon != 1.0 else dataset_config['lambda_recon']
            TRAIN_CONFIG['lambda_contrastive'] = args.lambda_contrastive if args.lambda_contrastive != 1.0 else dataset_config['lambda_contrastive']
            TRAIN_CONFIG['temperature'] = args.temperature
            TRAIN_CONFIG['score_weight_recon'] = args.score_weight_recon if args.score_weight_recon is not None else dataset_config['score_weight_recon']
            TRAIN_CONFIG['score_weight_consistency'] = args.score_weight_consistency if args.score_weight_consistency is not None else dataset_config['score_weight_consistency']
            TRAIN_CONFIG['batch_size'] = args.batch_size if args.batch_size is not None else dataset_config.get('batch_size', None)
        else:
            NUM_EPOCHS = args.epochs if args.epochs is not None else NUM_EPOCHS
            LEARNING_RATE = args.lr if args.lr is not None else LEARNING_RATE
            first_stage_epochs_arg = args.first_stage_epochs if args.first_stage_epochs is not None else None
            if first_stage_epochs_arg is not None:
                FIRST_STAGE_EPOCHS = int(first_stage_epochs_arg)
            SECOND_STAGE_EPOCHS = int(args.second_stage_epochs)
            GATE_LEARNING_RATE = float(args.gate_lr)
            TRAIN_CONFIG['lambda_recon'] = args.lambda_recon
            TRAIN_CONFIG['lambda_contrastive'] = args.lambda_contrastive
            TRAIN_CONFIG['temperature'] = args.temperature
            TRAIN_CONFIG['score_weight_recon'] = args.score_weight_recon if args.score_weight_recon is not None else 0.3
            TRAIN_CONFIG['score_weight_consistency'] = args.score_weight_consistency if args.score_weight_consistency is not None else 0.4
            TRAIN_CONFIG['batch_size'] = args.batch_size
        main(seed=args.seed)
